# test_data_manager.py - Test cases for DataManager class

import pytest
import os
import pandas as pd
import openpyxl
from datetime import datetime
import tempfile
import shutil

# Import your model classes
from models.student import Student
from models.room import Room
from models.exam_session import ExamSession
from models.data_manager import DataManager

class TestDataManager:
    @pytest.fixture
    def setup_files(self):
        """Create temporary test files"""
        # Create a temp directory
        temp_dir = tempfile.mkdtemp()
        
        # Create a test student data file
        student_file = os.path.join(temp_dir, "test_students.xlsx")
        
        # Create sample student data
        data = {
            'Hallticketno': ['21321A0401', '21321A0402', '21321A0501', '21321A0502'],
            'Year': [2, 2, 2, 2],
            'Semester': [1, 1, 1, 1],
            'Regulation': ['R21', 'R21', 'R21', 'R21']
        }
        df = pd.DataFrame(data)
        df.to_excel(student_file, index=False)
        
        # Create a test rooms file
        rooms_file = os.path.join(temp_dir, "test_rooms.xlsx")
        
        # Create sample room data
        data = {
            'Room No': ['101', '102'],
            'Rows': [3, 2],
            'Columns': [3, 2]
        }
        df = pd.DataFrame(data)
        df.to_excel(rooms_file, index=False)
        
        yield {
            'temp_dir': temp_dir,
            'student_file': student_file,
            'rooms_file': rooms_file
        }
        
        # Clean up
        shutil.rmtree(temp_dir)
    
    def test_load_student_data(self, setup_files):
        # Create DataManager instance
        data_manager = DataManager()
        
        # Load student data
        students = data_manager.load_student_data(setup_files['student_file'])
        
        # Check that we loaded the correct number of students
        assert len(students) == 4
        
        # Check that the students have the correct attributes
        assert students[0].hallticket_no == '21321A0401'
        assert students[0].year == 2
        assert students[0].semester == 1
        assert students[0].regulation == 'R21'
        assert students[0].branch_name == 'ECE'
        
        assert students[2].hallticket_no == '21321A0501'
        assert students[2].branch_name == 'CSE'
    
    def test_load_rooms(self, setup_files):
        # Create DataManager instance
        data_manager = DataManager()
        
        # Load room data
        rooms = data_manager.load_rooms(setup_files['rooms_file'])
        
        # Check that we loaded the correct number of rooms
        assert len(rooms) == 2
        
        # Check that the rooms have the correct attributes
        assert rooms[0].room_no == '101'
        assert rooms[0].rows == 3
        assert rooms[0].columns == 3
        assert rooms[0].capacity == 9
        
        assert rooms[1].room_no == '102'
        assert rooms[1].rows == 2
        assert rooms[1].columns == 2
        assert rooms[1].capacity == 4
    
    def test_create_empty_rooms_file(self, setup_files):
        # Create DataManager instance
        data_manager = DataManager()
        
        # Create an empty rooms file
        empty_file = os.path.join(setup_files['temp_dir'], "empty_rooms.xlsx")
        data_manager.create_empty_rooms_file(empty_file)
        
        # Check that the file exists
        assert os.path.exists(empty_file)
        
        # Load the file and check its structure
        wb = openpyxl.load_workbook(empty_file)
        ws = wb.active
        
        assert ws['A1'].value == "Room No"
        assert ws['B1'].value == "Rows"
        assert ws['C1'].value == "Columns"
    
    def test_save_rooms(self, setup_files):
        # Create DataManager instance
        data_manager = DataManager()
        
        # Create some rooms
        rooms = [
            Room("101", 3, 3),
            Room("102", 2, 2),
            Room("103", 4, 4)
        ]
        
        # Save the rooms
        save_file = os.path.join(setup_files['temp_dir'], "saved_rooms.xlsx")
        data_manager.save_rooms(rooms, save_file)
        
        # Check that the file exists
        assert os.path.exists(save_file)
        
        # Load the rooms and check they match
        loaded_rooms = data_manager.load_rooms(save_file)
        
        assert len(loaded_rooms) == 3
        assert loaded_rooms[0].room_no == "101"
        assert loaded_rooms[1].room_no == "102"
        assert loaded_rooms[2].room_no == "103"
    
    def test_save_seating_arrangement(self, setup_files):
        # Create DataManager instance
        data_manager = DataManager()
        
        # Create an exam session
        exam_session = ExamSession(
            exam_name="Test Exam",
            date=datetime(2025, 7, 10),
            year=2,
            semester=1
        )
        
        # Add a room
        room = Room("101", 2, 2)
        exam_session.add_room(room)
        
        # Add some students
        students = [
            Student("21321A0401"),
            Student("21321A0501"),
            Student("21321A0402"),
            Student("21321A0502")
        ]
        exam_session.add_students(students)
        
        # Assign students to seats
        room.set_seat(0, 0, students[0])
        room.set_seat(0, 1, students[1])
        room.set_seat(1, 0, students[2])
        room.set_seat(1, 1, students[3])
        
        # Save the seating arrangement
        seating_file = os.path.join(setup_files['temp_dir'], "seating.xlsx")
        data_manager.save_seating_arrangement(exam_session, seating_file)
        
        # Check that the file exists
        assert os.path.exists(seating_file)
        
        # Load the file and check its structure
        wb = openpyxl.load_workbook(seating_file)
        
        # Check that we have a sheet for each room
        assert "101" in wb.sheetnames
        
        # Check some basic structure in the sheet
        ws = wb["101"]
        assert "Mahaveer Institute of Science and Technology" in ws['A1'].value
        assert "Seating Arrangement" in ws['A2'].value
        assert "ROOM NO:" in ws['A3'].value
        assert "DT:" in ws.cell(row=3, column=2).value