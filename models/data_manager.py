import os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from models.student import Student
from models.room import Room
from models.exam_session import ExamSession
from utils.constants import BRANCH_CODES


class DataManager:
    """
    DataManager class for handling data operations
    """
    def __init__(self):
        """Initialize the DataManager"""
        pass
    
    def load_student_data(self, file_path):
        """
        Load student data from Excel file
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            list: List of Student objects
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file format is invalid
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            # Read all sheets from the Excel file
            df_list = []
            for sheet_name in pd.read_excel(file_path, sheet_name=None):
                df_list.append(pd.read_excel(file_path, sheet_name=sheet_name))
            
            # Concatenate all dataframes
            student_df = pd.concat(df_list)
            
            # Create Student objects from dataframe
            students = []
            for _, row in student_df.iterrows():
                hallticket_no = row.get('Hallticketno', '')
                year = row.get('Year', None)
                semester = row.get('Semester', None)
                regulation = row.get('Regulation', None)
                
                student = Student(
                    hallticket_no=hallticket_no,
                    year=year,
                    semester=semester,
                    regulation=regulation
                )
                students.append(student)
            
            return students
            
        except Exception as e:
            raise ValueError(f"Error loading student data: {str(e)}")
    
    def load_rooms(self, file_path='rooms.xlsx'):
        """
        Load room data from Excel file
        
        Args:
            file_path (str, optional): Path to the Excel file (default: 'rooms.xlsx')
            
        Returns:
            list: List of Room objects
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file format is invalid
        """
        if not os.path.exists(file_path):
            # Create empty rooms file if it doesn't exist
            self.create_empty_rooms_file(file_path)
        
        try:
            # Read rooms data from Excel file
            rooms_df = pd.read_excel(file_path)
            
            # Create Room objects from dataframe
            rooms = []
            for _, row in rooms_df.iterrows():
                room_no = row.get('Room No', '')
                rows = row.get('Rows', 0)
                columns = row.get('Columns', 0)
                
                room = Room(room_no=room_no, rows=rows, columns=columns)
                rooms.append(room)
            
            return rooms
            
        except Exception as e:
            raise ValueError(f"Error loading room data: {str(e)}")
    
    def create_empty_rooms_file(self, file_path='rooms.xlsx'):
        """
        Create an empty rooms Excel file with headers
        
        Args:
            file_path (str, optional): Path for the Excel file (default: 'rooms.xlsx')
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Room No", "Rows", "Columns"])
        workbook.save(file_path)
    
    def save_rooms(self, rooms, file_path='rooms.xlsx'):
        """
        Save room data to Excel file
        
        Args:
            rooms (list): List of Room objects
            file_path (str, optional): Path for the Excel file (default: 'rooms.xlsx')
            
        Raises:
            ValueError: If there is an error saving the data
        """
        try:
            # Create dataframe from Room objects
            data = []
            for room in rooms:
                data.append({
                    'Room No': room.room_no,
                    'Rows': room.rows,
                    'Columns': room.columns
                })
            
            rooms_df = pd.DataFrame(data)
            
            # Save dataframe to Excel file
            rooms_df.to_excel(file_path, index=False)
            
        except Exception as e:
            raise ValueError(f"Error saving room data: {str(e)}")
    
    def save_seating_arrangement(self, exam_session, file_path=None):
        """
        Save seating arrangement to Excel file
        
        Args:
            exam_session (ExamSession): Exam session with seating arrangement
            file_path (str, optional): Path for the Excel file
                (default: generated based on date)
                
        Returns:
            str: Path to the saved file
            
        Raises:
            ValueError: If there is an error saving the data
        """
        if file_path is None:
            # Generate file name based on date - IMPORTANT: Match original filename format
            date_str = exam_session.get_formatted_date("%d_%b_%y")
            file_path = f"{date_str}.xlsx"  # Match original naming convention
        
        try:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # Create a sheet for each room
            for room in exam_session.rooms:
                self._create_room_sheet(workbook, room, exam_session)
            
            # Save the workbook
            workbook.save(file_path)
            return file_path
            
        except Exception as e:
            raise ValueError(f"Error saving seating arrangement: {str(e)}")
    
    def _create_room_sheet(self, workbook, room, exam_session):
        """
        Create a worksheet for a room's seating arrangement
        
        Args:
            workbook (Workbook): Excel workbook
            room (Room): Room object
            exam_session (ExamSession): Exam session data
        """
        # Create a sheet with room number as name
        sheet_name = str(room.room_no)
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Styling constants
        border_style = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal="center", vertical="center")
        title_font = Font(size=24, name='Times New Roman')
        yellow_fill = PatternFill(fill_type="solid", patternType='solid', start_color="FFFFFF00")
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 15  # For ENTRANCE
        for i in range(room.columns):
            col_letter = get_column_letter(i + 2)  # Starting from column B
            worksheet.column_dimensions[col_letter].width = 15
        
        # Calculate the last column letter
        last_col = get_column_letter(room.columns + 1)  # +1 for ENTRANCE column
        
        # Header: Institution name
        worksheet.merge_cells(f'A1:{last_col}1')
        cell = worksheet['A1']
        cell.value = 'Mahaveer Institute of Science and Technology'
        cell.font = title_font
        cell.alignment = alignment
        cell.fill = yellow_fill
        cell.border = border_style
        
        # Header: Seating arrangement
        worksheet.merge_cells(f'A2:{last_col}2')
        cell = worksheet['A2']
        title_text = 'Seating Arrangement'
        if exam_session.year and exam_session.semester:
            title_text += f" {exam_session.year}-{exam_session.semester}"
        title_text += f" External Examination {exam_session.get_formatted_date('%b %Y')}"
        cell.value = title_text
        cell.font = title_font
        cell.alignment = alignment
        cell.fill = yellow_fill
        cell.border = border_style
        
        # Room number (left side)
        mid_col = (room.columns + 1) // 2  # Include ENTRANCE column in the calculation
        mid_col_letter = get_column_letter(mid_col)
        worksheet.merge_cells(f'A3:{mid_col_letter}3')
        cell = worksheet['A3']
        cell.value = f"ROOM NO: {room.room_no}"  # Include the actual room number
        cell.font = title_font
        cell.alignment = alignment
        cell.border = border_style
        
        # Date (right side)
        mid_col_plus_one = get_column_letter(mid_col + 1)
        worksheet.merge_cells(f'{mid_col_plus_one}3:{last_col}3')
        cell = worksheet[f'{mid_col_plus_one}3']
        cell.value = f"DT: {exam_session.get_formatted_date('%d-%b-%Y')}"
        cell.font = title_font
        cell.alignment = alignment
        cell.border = border_style
        
        # Strength label
        worksheet.merge_cells(f'A4:{last_col}4')
        cell = worksheet['A4']
        cell.value = "Strength:"
        cell.font = title_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border_style
        
        # Calculate branch-wise strength for each column
        column_branches = {}
        for c in range(room.columns):
            branch_counts = {}
            for r in range(room.rows):
                student = room.get_seat(r, c)
                if student and student.branch_name:
                    branch = student.branch_name
                    if branch not in branch_counts:
                        branch_counts[branch] = 0
                    branch_counts[branch] += 1
            
            # Determine the predominant branch for this column
            if branch_counts:
                # Sort branches by count (descending) and then by name
                sorted_branches = sorted(branch_counts.items(), key=lambda x: (-x[1], x[0]))
                # Use the most common branch as the label
                column_branches[c] = sorted_branches[0][0]
            else:
                column_branches[c] = "EMPTY"
        
        # Add branch name headers
        cell = worksheet['A5']  # Empty cell in top-left of the table
        cell.border = border_style
        
        for i in range(room.columns):
            col = get_column_letter(i + 2)  # Start from column B
            cell = worksheet[f'{col}5']
            cell.value = column_branches[i] if i in column_branches else f"COL {i + 1}"
            cell.font = title_font
            cell.alignment = alignment
            cell.border = border_style
        
        # Add ENTRANCE label
        cell = worksheet['A6']
        cell.value = 'ENTRANCE'
        cell.alignment = alignment
        cell.border = border_style
        
        # Add seating arrangement
        start_row = 6  # Row for first student
        for r in range(room.rows):
            for c in range(room.columns):
                student = room.get_seat(r, c)
                # Column is c+2 because column A is for ENTRANCE
                cell = worksheet.cell(row=start_row + r, column=c + 2)
                if student:
                    cell.value = student.hallticket_no
                else:
                    cell.value = None  # Leave empty instead of '---'
                cell.alignment = alignment
                cell.border = border_style
        
        # Set row heights
        for row in range(1, start_row + room.rows):
            if row <= 4:  # Headers
                worksheet.row_dimensions[row].height = 30
            else:  # Student rows
                worksheet.row_dimensions[row].height = 25
        
        # Apply borders to all cells in the table area
        for row in range(1, start_row + room.rows):
            for col in range(1, room.columns + 2):  # +2: one for ENTRANCE, one to convert to 1-based
                cell = worksheet.cell(row=row, column=col)
                cell.border = border_style
        
        # Set print settings
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(
            worksheet, paper_size=4, orientation='landscape'
        )
    def save_summary(self, exam_session, file_path=None):
        """
        Save summary of students by branch and regulation
        
        Args:
            exam_session (ExamSession): Exam session data
            file_path (str, optional): Path for the Excel file
                (default: generated based on date)
                
        Returns:
            str: Path to the saved file
            
        Raises:
            ValueError: If there is an error saving the data
        """
        if file_path is None:
            # Generate file name based on date
            date_str = exam_session.get_formatted_date("%d_%b_%y")
            file_path = f"{date_str}_summary.xlsx"
        
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            # Add headers
            worksheet['A1'] = 'Branch Name'
            worksheet['B1'] = 'Regulation'
            worksheet['C1'] = 'Total Students'
            
            # Count students by branch and regulation
            counts = {}
            for student in exam_session.students:
                key = (student.branch_name, student.regulation)
                if key not in counts:
                    counts[key] = 0
                counts[key] += 1
            
            # Add data rows
            row_num = 2
            for (branch_name, regulation), count in counts.items():
                worksheet.cell(row=row_num, column=1, value=branch_name)
                worksheet.cell(row=row_num, column=2, value=regulation)
                worksheet.cell(row=row_num, column=3, value=count)
                row_num += 1
            
            # Save the workbook
            workbook.save(file_path)
            return file_path
            
        except Exception as e:
            raise ValueError(f"Error saving summary: {str(e)}")
    
    def save_out_sheet(self, exam_session, file_path=None):
        """
        Save out sheet with room-wise breakdown of students
        
        Args:
            exam_session (ExamSession): Exam session data
            file_path (str, optional): Path for the Excel file
                (default: generated based on date)
                
        Returns:
            str: Path to the saved file
            
        Raises:
            ValueError: If there is an error saving the data
        """
        if file_path is None:
            # Generate file name based on date
            date_str = exam_session.get_formatted_date("%d_%b_%y")
            file_path = f"{date_str}_out_sheet.xlsx"
        
        try:
            # Get room summary data
            room_data = self._get_room_summary_data(exam_session)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # Get date components
            date_now = datetime.now()
            date = date_now.strftime("%d")
            mon = date_now.strftime("%b")
            year = date_now.strftime("%y")
            
            # Create worksheet
            worksheet = workbook.create_sheet(f"{date}_{mon}_{year}_out_sheet")
            
            # Styling constants
            black = "FF000000"
            yellow = "FFFFFF00"
            thin = Side(style="thin", color=black)
            alignment = Alignment(horizontal="center", vertical="center")
            border_style = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill_style = PatternFill(fill_type="solid", patternType='solid', start_color=yellow)
            
            # Set column widths
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['C'].width = 7
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 7
            
            # Add headers
            worksheet.merge_cells('A2:H2')
            worksheet['A2'].font = Font(size=23, name='Times New Roman')
            worksheet['A2'].alignment = alignment
            worksheet['A2'] = 'Mahaveer Institute of Science and Technology'
            worksheet['A2'].fill = fill_style
            worksheet['A2'].border = border_style
            
            worksheet.merge_cells('A3:H3')
            worksheet['A3'].font = Font(size=23, name='Times New Roman')
            worksheet['A3'].alignment = alignment
            worksheet['A3'] = 'NEW BLOCK'
            worksheet['A3'].fill = fill_style
            worksheet['A3'].border = border_style
            
            worksheet.merge_cells('A4:H4')
            worksheet['A4'].font = Font(size=23, name='Times New Roman')
            worksheet['A4'].alignment = Alignment(horizontal='right')
            worksheet['A4'] = f"Dt : {date}.{mon}.{year}"
            worksheet['A4'].fill = fill_style
            worksheet['A4'].border = border_style
            
            worksheet.merge_cells('A5:H5')
            worksheet['A5'].font = Font(size=23, name='Times New Roman')
            worksheet['A5'].alignment = alignment
            worksheet['A5'] = f"JNTU External Examination- {mon}- {year}"
            worksheet['A5'].fill = fill_style
            worksheet['A5'].border = border_style
            
            worksheet.merge_cells('A6:H6')
            worksheet['A6'].font = Font(size=23, name='Times New Roman')
            worksheet['A6'].alignment = alignment
            worksheet['A6'] = "Seating Arrangement"
            worksheet['A6'].fill = fill_style
            worksheet['A6'].border = border_style
            
            worksheet.merge_cells('A7:H7')
            
            # Add year and semester if available
            if exam_session.students and exam_session.students[0].year and exam_session.students[0].semester:
                year_val = exam_session.students[0].year
                semester_val = exam_session.students[0].semester
                worksheet['A7'] = f"B.TECH {year_val}-{semester_val} Reg/Supply Room Numbers"
            else:
                worksheet['A7'] = "B.TECH Reg/Supply Room Numbers"
                
            worksheet['A7'].font = Font(size=23, name='Times New Roman')
            worksheet['A7'].alignment = alignment
            worksheet['A7'].fill = fill_style
            worksheet['A7'].border = border_style
            
            # Column headers
            worksheet['A8'] = 'Room No'
            worksheet['A8'].alignment = alignment
            worksheet['B8'] = 'College Code'
            worksheet['B8'].alignment = alignment
            worksheet['C8'] = 'Branch'
            worksheet['C8'].alignment = alignment
            worksheet.merge_cells('D8:G8')
            worksheet['D8'] = 'H.T.Nos'
            worksheet['D8'].alignment = alignment
            worksheet['H8'] = 'Total'
            worksheet['H8'].alignment = alignment
            
            # Merge cells for HT numbers
            for row in range(9, 33):
                worksheet.merge_cells(f'D{row}:G{row}')
            
            # Set row heights
            for row in range(8, 33):
                worksheet.row_dimensions[row].height = 75
            
            # Add data
            current_row = 9
            
            # Process room data
            for room_no, room_info in room_data.items():
                branch_data = room_info['branches']
                college_codes = room_info['college_codes']
                
                # Room number
                first_row = current_row
                last_row = current_row + len(branch_data) - 1
                
                if last_row >= first_row:
                    worksheet.merge_cells(start_row=first_row, start_column=1, 
                                         end_row=last_row, end_column=1)
                    worksheet.cell(row=first_row, column=1).value = room_no
                
                # Branch, college codes, and halltickets
                for branch_name, halltickets in branch_data.items():
                    # Branch
                    worksheet.cell(row=current_row, column=3).value = branch_name
                    
                    # College codes
                    if branch_name in college_codes:
                        codes_str = ','.join(college_codes[branch_name])
                        worksheet.cell(row=current_row, column=2).value = codes_str
                    
                    # Halltickets
                    halltickets_str = ','.join(halltickets)
                    cell = worksheet.cell(row=current_row, column=4)
                    cell.value = halltickets_str
                    cell.alignment = Alignment(wrap_text=True)
                    
                    # Total
                    worksheet.cell(row=current_row, column=8).value = len(halltickets)
                    
                    current_row += 1
            
            # Apply borders
            for row in worksheet['A2':'H32']:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = alignment
            
            # Set print settings
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(
                worksheet, paper_size=4, orientation='portrait'
            )
            
            # Save workbook
            workbook.save(file_path)
            return file_path
            
        except Exception as e:
            raise ValueError(f"Error saving out sheet: {str(e)}")
    
    def _get_room_summary_data(self, exam_session):
        """
        Calculate summary data for each room
        
        Args:
            exam_session (ExamSession): Exam session data
            
        Returns:
            dict: Summary data for rooms
        """
        room_data = {}
        
        for room in exam_session.rooms:
            branch_data = {}
            college_code_data = {}
            
            # Iterate through each seat
            for r in range(room.rows):
                for c in range(room.columns):
                    student = room.get_seat(r, c)
                    if student and student.hallticket_no:
                        # Count by branch
                        if student.branch_name not in branch_data:
                            branch_data[student.branch_name] = []
                        branch_data[student.branch_name].append(student.hallticket_no)
                        
                        # Track college codes
                        college_code = student.get_college_code()
                        if student.branch_name not in college_code_data:
                            college_code_data[student.branch_name] = []
                        if college_code and college_code not in college_code_data[student.branch_name]:
                            college_code_data[student.branch_name].append(college_code)
            
            # Remove empty branches
            empty_branches = [branch for branch in branch_data if len(branch_data[branch]) == 0]
            for branch in empty_branches:
                del branch_data[branch]
                if branch in college_code_data:
                    del college_code_data[branch]
            
            room_data[room.room_no] = {
                'branches': branch_data,
                'college_codes': college_code_data
            }
        
        return room_data